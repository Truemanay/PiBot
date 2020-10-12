from openpyxl import *
from tkinter import *

# This will open an excel sheet
wb = load_workbook('C:\\Users\\uC244458\\RADZs\\PythonStuff.xlsx')

# This corresponds with the actual excel sheet
sheet = wb.active

def excel():

    # excel sheet columns
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Password"

def focus1(event):
    name_field.focus_set()

def focus2(event):
    password_field.focus_set()

def clear():
    name_field.delete(0, END)
    password_field.delete(0, END)

# This takes the GUI data and puts in into the excel sheet
def insert():

    current_row = sheet.max_row
    current_column = sheet.max_column
    sheet.cell(row=current_row + 1, column=1).value = name_field.get()
    sheet.cell(row=current_row + 1, column=2).value = password_field.get()

    # save the file
    wb.save('C:\\Users\\uC244458\\RADZs\\PythonStuff.xlsx')

    name_field.focus_set()
    #password_field.focus_set()
    clear()


# DRIVER
if __name__ == "__main__":

    # creates the GUI
    root = Tk()
    root.configure(background='light blue')
    root.title("WebScraper UI Project")
    root.geometry("450x200")
    excel()
    heading = Label(root, text="Login To Our Project", bg="light blue")
    name = Label(root, text="Name", bg="light blue")
    password = Label(root, text="Password", bg="light blue")

    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    password.grid(row=2, column=0)

    name_field = Entry(root)
    password_field = Entry(root)

    #binding methods?
    name_field.bind("<Return>", focus1)
    password_field.bind("<Return>", focus2)

    name_field.grid(row=1, column=1, ipadx="100")
    password_field.grid(row=2, column=1, ipadx="100")

    #excel function call
    excel()

    #submit Button
    submit = Button(root, text="Login", fg="Black", bg="MistyRose2", command=insert)
    submit.grid(row=8, column=1)

    #Used to actually start the code
    root.mainloop()
